

from NewApi.common import project_path
from NewApi.common.read_config import ReadConfig
from openpyxl import load_workbook


class DoExcel(object):
    '''完成测试用例的读取以及测试结果的写入'''

    def __init__(self, file_name, file_sheet):
        self.file_name = file_name
        self.file_sheet = file_sheet

    def read_data(self,section):
        '''从excel中读取数据，返回，一次性读取效率高
        session:配置文件的片段名，用于指定读取的数据'''
        case_id = ReadConfig(project_path.conf_path).get_data(section,"case_id")
        wb = load_workbook(self.file_name)
        sheet = wb[self.file_sheet]

        tel = self.get_tel()

        #读取格式的区别，这里不利于迭代数据，存入列表中，其中各个数据用字典表示
        test_data = []
        for i in range(2, sheet.max_row+1):
            row_data = {}
            row_data["CaseId"] = sheet.cell(i, 1).value
            row_data["Module"] = sheet.cell(i, 2).value
            row_data["Title"] = sheet.cell(i, 3).value
            row_data["Url"] = sheet.cell(i, 4).value
            row_data["Method"] = sheet.cell(i, 5).value

            #拿到表格中的tel值进行替换,并且赋值给回去给row["Params"]
            #这里使用了str中find方法，返回第一次出现的下标，没有返回-1
            if sheet.cell(i,6).value.find("tel") != -1:
                row_data["Params"] = sheet.cell(i,6).value.replace("tel",str(tel))
                self.update_tel(int(tel)+1)
               # print(row_data["Params"])
            #如果不存在tel这个值就说明显示的是手机号，这里就不需要替换了
            else:
                row_data['Params'] = sheet.cell(i, 6).value

            #用于处理新增的sql语句
            row_data["Sql"] = sheet.cell(i,7).value
            row_data["ExceptedResult"] = sheet.cell(i, 8).value
            test_data.append(row_data)

        wb.close()

        #根据配置返回数据,规定如果是all就返回所有的数据，否则按照列表来显示
        final_data = []
        if case_id == "all":
            final_data = test_data
        else:
            for i in case_id:
                final_data.append(test_data[i-1])
        return final_data


    def write_data(self,row,colum,data):
        '''根据定位的写回数据表中，'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.file_sheet]
        sheet.cell(row,colum,data)
        wb.save(self.file_name)
        wb.close()

    def get_tel(self):
        '''获取存在Excel里面的tel表单的数值'''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        wb.close()
        return sheet.cell(1,2).value

    def update_tel(self,new_tel):
        '''写回表格中tel的数值，把手机号+1'''
        wb = load_workbook(self.file_name)
        sheet=wb['tel']
        sheet.cell(1,2,new_tel)
        wb.save(self.file_name)
        wb.close()



if __name__ == '__main__':
    sheet_name = "register"
    print(DoExcel(project_path.case_path,sheet_name).read_data("RechargeCase"))